#!/usr/bin/env python3
"""
RetireFlow Dashboard Generator
讀取 RetireFlow_DB.xlsx → 產生 index.html
用法：python scripts/generate.py [xlsx路徑]
"""
import sys
import json
import math
from datetime import datetime
from pathlib import Path

# pip install openpyxl --break-system-packages
from openpyxl import load_workbook

# ── 設定 ──────────────────────────────────────────────────────
CONFIG = {
    "SHEET_HOLDINGS" : "工作表1",
    "SHEET_FUNDS"    : "基金帳戶",
    "SHEET_DEBT"     : "負債清單",
    "SHEET_HISTORY"  : "資產歷史紀錄",
    "RETIRE_TARGET"  : 160_000_000,   # 16,000萬
    "RETIRE_YEAR"    : 2028,
    "CURRENT_AGE"    : 52,
    "ALLOC_TW"       : 0.40,
    "ALLOC_US"       : 0.30,
    "ALLOC_JP"       : 0.10,
    "ALLOC_FUND"     : 0.10,
    "ALLOC_CASH"     : 0.10,
}

# ── 工具函式 ──────────────────────────────────────────────────
def w(n):  return round(n / 10000)
def p(n):  return f"{n:.1f}"
def s(n):  return f"+{n:.1f}" if n >= 0 else f"{n:.1f}"
def fmt_date(val):
    if isinstance(val, datetime): return val.strftime("%m/%d")
    return str(val)[5:].replace("-", "/") if val else ""
def annualized(pct, days):
    try: return ((1 + pct/100) ** (365/days) - 1) * 100
    except: return 0

# ── 讀取資料 ──────────────────────────────────────────────────
def read_data(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # 歷史紀錄
    ws = wb[CONFIG["SHEET_HISTORY"]]
    history = []
    for row in ws.iter_rows(values_only=True):
        if not row[0] or row[0] == "紀錄日期": continue
        history.append({
            "date": fmt_date(row[0]),
            "total": float(row[1] or 0), "debt": float(row[2] or 0),
            "net":   float(row[3] or 0), "div":  float(row[4] or 0),
            "tw":    float(row[5] or 0), "us":   float(row[6] or 0),
            "jp":    float(row[7] or 0), "fund": float(row[8] or 0),
        })

    # 負債
    ws2 = wb[CONFIG["SHEET_DEBT"]]
    debts = []
    for row in ws2.iter_rows(values_only=True):
        if not row[0] or row[0] == "負債項目(如房貸,質借)": continue
        debts.append({"name": row[0], "bank": row[1], "amount": float(row[2] or 0), "rate": float(row[3] or 0)})

    # 基金
    ws3 = wb[CONFIG["SHEET_FUNDS"]]
    funds = []
    for row in ws3.iter_rows(values_only=True):
        if not row[0] or row[0] == "基金名稱": continue
        funds.append({"name": row[0], "platform": row[1], "amount": float(row[2] or 0), "yld": float(row[3] or 0)})

    # 持股
    ws4 = wb[CONFIG["SHEET_HOLDINGS"]]
    holdings = {"tw": [], "us": [], "jp": []}
    for row in ws4.iter_rows(values_only=True):
        if not row[0] or row[0] == "市場" or not row[2]: continue
        item = {"market": row[0], "broker": str(row[1] or ""), "code": str(row[2]),
                "shares": float(row[3] or 0), "yld": float(row[4] or 0), "name": str(row[5] or row[2])}
        if row[0] == "台股": holdings["tw"].append(item)
        elif row[0] == "美股": holdings["us"].append(item)
        elif row[0] == "日股": holdings["jp"].append(item)

    wb.close()
    return history, debts, funds, holdings

# ── 計算指標 ──────────────────────────────────────────────────
def calc_metrics(history, debts, funds, holdings):
    L = history[-1]; P = history[-2] if len(history) > 1 else L; F = history[0]
    days = max(len(history), 1)

    total_debt = sum(d["amount"] for d in debts)
    annual_int = sum(d["amount"] * d["rate"] for d in debts)
    fund_total = sum(f["amount"] for f in funds)

    day_chg     = L["total"] - P["total"]
    day_chg_pct = day_chg / P["total"] * 100 if P["total"] else 0
    period_chg  = L["total"] - F["total"]
    period_pct  = period_chg / F["total"] * 100 if F["total"] else 0
    achieve     = L["total"] / CONFIG["RETIRE_TARGET"] * 100
    tw_achieve  = L["tw"]   / (CONFIG["RETIRE_TARGET"] * CONFIG["ALLOC_TW"])   * 100
    us_achieve  = L["us"]   / (CONFIG["RETIRE_TARGET"] * CONFIG["ALLOC_US"])   * 100
    jp_achieve  = L["jp"]   / (CONFIG["RETIRE_TARGET"] * CONFIG["ALLOC_JP"])   * 100
    fund_ach    = L["fund"] / (CONFIG["RETIRE_TARGET"] * CONFIG["ALLOC_FUND"]) * 100

    tw_ret  = (L["tw"]  - F["tw"])  / F["tw"]  * 100 if F["tw"]  else 0
    us_ret  = (L["us"]  - F["us"])  / F["us"]  * 100 if F["us"]  else 0
    jp_ret  = (L["jp"]  - F["jp"])  / F["jp"]  * 100 if F["jp"]  else 0
    fund_ret= (L["fund"]- F["fund"])/ F["fund"] * 100 if F["fund"] else 0

    tw_pct   = L["tw"]  / L["total"] * 100 if L["total"] else 0
    us_pct   = L["us"]  / L["total"] * 100 if L["total"] else 0
    jp_pct   = L["jp"]  / L["total"] * 100 if L["total"] else 0
    fund_pct = L["fund"]/ L["total"] * 100 if L["total"] else 0
    debt_ratio = total_debt / L["total"] * 100 if L["total"] else 0

    now = datetime.now()
    retire_date = datetime(CONFIG["RETIRE_YEAR"], 12, 1)
    diff_months = max(0, int((retire_date - now).days / 30.44))

    return {
        "L": L, "P": P, "F": F, "history": history,
        "debts": debts, "funds": funds, "holdings": holdings,
        "total_debt": total_debt, "annual_int": annual_int, "fund_total": fund_total,
        "day_chg": day_chg, "day_chg_pct": day_chg_pct,
        "period_chg": period_chg, "period_pct": period_pct, "period_days": days,
        "achieve": achieve, "tw_achieve": tw_achieve, "us_achieve": us_achieve,
        "jp_achieve": jp_achieve, "fund_achieve": fund_ach,
        "tw_ret": tw_ret, "us_ret": us_ret, "jp_ret": jp_ret, "fund_ret": fund_ret,
        "tw_ann": annualized(tw_ret, days), "us_ann": annualized(us_ret, days),
        "jp_ann": annualized(jp_ret, days), "fund_ann": annualized(fund_ret, days),
        "tw_pct": tw_pct, "us_pct": us_pct, "jp_pct": jp_pct, "fund_pct": fund_pct,
        "debt_ratio": debt_ratio,
        "retire_yrs": diff_months // 12, "retire_mths": diff_months % 12,
        "update_time": now.strftime("%Y/%m/%d %H:%M"),
    }

# ── 產生 HTML ─────────────────────────────────────────────────
def build_html(m):
    L = m["L"]; F = m["F"]
    total = w(L["total"]); debt = w(m["total_debt"]); net = w(L["net"]); div = w(L["div"])
    tw = w(L["tw"]); us = w(L["us"]); jp = w(L["jp"]); fund = w(L["fund"])
    int_w = w(m["annual_int"]); div_net = div - int_w
    need  = w(CONFIG["RETIRE_TARGET"] - L["total"])
    day_str = ("↑ +" if m["day_chg"] >= 0 else "↓ ") + str(abs(w(m["day_chg"]))) + f'萬 ({s(m["day_chg_pct"])}%)'
    peri_str = s(w(m["period_chg"])) + f'萬（{s(m["period_pct"])}%）'

    # 歷史 JSON
    h_dates = json.dumps([h["date"] for h in m["history"]])
    h_total = json.dumps([w(h["total"]) for h in m["history"]])
    h_net   = json.dumps([w(h["net"])   for h in m["history"]])
    h_debt  = json.dumps([w(h["debt"])  for h in m["history"]])
    h_tw    = json.dumps([w(h["tw"])    for h in m["history"]])
    h_us    = json.dumps([w(h["us"])    for h in m["history"]])
    h_jp    = json.dumps([w(h["jp"])    for h in m["history"]])
    h_fund  = json.dumps([w(h["fund"])  for h in m["history"]])

    # 持股 HTML
    tw_rows = "".join(
        f'<div class="row"><div><div class="row-name">{h["code"]} {h["name"]}</div>'
        f'<div class="row-tag">{h["broker"]}｜{int(h["shares"]):,}股</div></div>'
        f'<div style="font-size:11px" class="{"up" if h["yld"]>0 else ""}">'
        f'{"殖利率"+str(h["yld"])+"%" if h["yld"]>0 else "—"}</div></div>'
        for h in m["holdings"]["tw"] if h["shares"] > 0
    )
    us_rows = "".join(
        f'<div class="row"><div><div class="row-name">{h["code"]}</div>'
        f'<div class="row-tag">{h["broker"]}｜{h["shares"]}股</div></div></div>'
        for h in m["holdings"]["us"] if h["shares"] > 0
    )
    jp_rows = "".join(
        f'<div class="row"><div><div class="row-name">{h["code"]} {h["name"]}</div>'
        f'<div class="row-tag">{h["broker"]}｜{int(h["shares"]):,}股</div></div></div>'
        for h in m["holdings"]["jp"] if h["shares"] > 0
    )
    debt_rows = "".join(
        f'<div class="row"><span>{d["name"]}（{d["bank"]}）</span>'
        f'<span><b class="dn">{w(d["amount"]):,}萬</b> '
        f'<span style="font-size:11px;color:var(--hint)">@{d["rate"]*100:.1f}%</span></span></div>'
        for d in m["debts"]
    )
    fund_rows = "".join(
        f'<div class="row"><span>{f["name"]}（{f["platform"]}）</span>'
        f'<span>{w(f["amount"]):,}萬 <span style="font-size:11px;color:var(--hint)">殖利率{f["yld"]}%</span></span></div>'
        for f in m["funds"]
    )

    RT = CONFIG["RETIRE_TARGET"]
    ach = p(m["achieve"])
    tw_p = p(m["tw_pct"]); us_p = p(m["us_pct"]); jp_p = p(m["jp_pct"]); fund_p = p(m["fund_pct"])
    tw_a = p(m["tw_achieve"]); us_a = p(m["us_achieve"]); jp_a = p(m["jp_achieve"]); fund_a = p(m["fund_achieve"])
    tw_ach_diff = s(m["tw_pct"] - CONFIG["ALLOC_TW"]*100)
    us_ach_diff = s(m["us_pct"] - CONFIG["ALLOC_US"]*100)
    jp_ach_diff = s(m["jp_pct"] - CONFIG["ALLOC_JP"]*100)
    fd_ach_diff = s(m["fund_pct"] - CONFIG["ALLOC_FUND"]*100)
    tw_ach_color = "#E24B4A" if m["tw_pct"] > 45 else "#1D9E75"

    html = f"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="default">
<meta name="apple-mobile-web-app-title" content="RetireFlow">
<title>RetireFlow 退休財務儀表板</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
:root{{--bg:#f5f4f0;--card:#fff;--border:#e8e6df;--text:#1a1a1a;--muted:#666;--hint:#999;--blue:#378ADD;--green:#1D9E75;--amber:#EF9F27;--red:#E24B4A;--purple:#7F77DD;--radius:12px;--radius-sm:8px}}
@media(prefers-color-scheme:dark){{:root{{--bg:#1a1a18;--card:#252523;--border:#333;--text:#f0efe9;--muted:#aaa;--hint:#555}}}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.5;min-height:100vh}}
.wrap{{max-width:980px;margin:0 auto;padding:16px}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:18px;flex-wrap:wrap;gap:10px}}
.header h1{{font-size:20px;font-weight:600}}.header p{{font-size:12px;color:var(--muted);margin-top:2px}}
.pill{{display:flex;align-items:center;gap:6px;padding:5px 12px;border-radius:99px;border:1px solid var(--border);font-size:12px;color:var(--muted);background:var(--card)}}
.dot{{width:9px;height:9px;border-radius:50%;display:inline-block;flex-shrink:0}}
.dot-green{{background:#1D9E75;box-shadow:0 0 0 3px #E1F5EE}}.dot-amber{{background:#EF9F27;box-shadow:0 0 0 3px #FAEEDA}}.dot-red{{background:#E24B4A;box-shadow:0 0 0 3px #FCEBEB}}
.g4{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:12px}}
.g3{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin-bottom:12px}}
.g2{{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin-bottom:12px}}
@media(max-width:680px){{.g4{{grid-template-columns:repeat(2,1fr)}}.g3{{grid-template-columns:repeat(2,1fr)}}.g2{{grid-template-columns:1fr}}}}
@media(max-width:380px){{.g4{{grid-template-columns:1fr}}}}
.card{{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);padding:14px 16px}}
.metric{{background:var(--bg);border-radius:var(--radius-sm);padding:12px 14px}}
.metric-label{{font-size:10px;color:var(--muted);margin-bottom:4px;font-weight:600;letter-spacing:0.06em;text-transform:uppercase}}
.metric-value{{font-size:22px;font-weight:600;letter-spacing:-0.5px;line-height:1.1}}
.metric-sub{{font-size:11px;color:var(--hint);margin-top:3px}}
.up{{color:#1D9E75}}.dn{{color:#E24B4A}}
.section-title{{font-size:10px;font-weight:600;letter-spacing:0.07em;text-transform:uppercase;color:var(--hint);margin-bottom:12px}}
.mb12{{margin-bottom:12px}}
.prog-bg{{height:8px;background:var(--bg);border-radius:99px;overflow:hidden;margin:6px 0 3px}}
.prog-fill{{height:100%;border-radius:99px}}
.tab-bar{{display:flex;gap:5px;margin-bottom:14px;flex-wrap:wrap}}
.tab{{padding:6px 14px;border-radius:99px;font-size:12px;cursor:pointer;border:1px solid var(--border);background:transparent;color:var(--muted);font-family:inherit;transition:all .15s}}
.tab:hover{{background:var(--bg)}}.tab.active{{background:var(--text);color:var(--card);border-color:transparent}}
.tab-content{{display:none}}.tab-content.active{{display:block}}
.row{{display:flex;justify-content:space-between;align-items:center;padding:7px 0;border-bottom:1px solid var(--border);font-size:13px}}
.row:last-child{{border-bottom:none}}
.row-name{{font-weight:500}}.row-tag{{font-size:11px;color:var(--hint)}}
.alert{{border-left:3px solid;border-radius:0 var(--radius-sm) var(--radius-sm) 0;padding:9px 12px;margin-bottom:8px;font-size:12px;line-height:1.6}}
.alert-warn{{border-color:#EF9F27;background:#FAEEDA;color:#633806}}
.alert-good{{border-color:#1D9E75;background:#E1F5EE;color:#085041}}
.alert-danger{{border-color:#E24B4A;background:#FCEBEB;color:#501313}}
.alert-info{{border-color:#378ADD;background:#E6F1FB;color:#0C447C}}
.vix-bar{{height:7px;border-radius:4px;background:linear-gradient(to right,#1D9E75,#EF9F27,#E24B4A);margin:8px 0 4px;position:relative}}
.vix-pin{{position:absolute;top:-4px;width:14px;height:14px;border-radius:50%;background:var(--text);border:2px solid var(--card);transform:translateX(-50%)}}
.risk-row{{display:flex;align-items:center;gap:10px;margin-bottom:8px}}
.risk-label{{font-size:12px;width:80px;flex-shrink:0;color:var(--muted)}}
.risk-bar-bg{{flex:1;height:7px;background:var(--bg);border-radius:99px;overflow:hidden}}
.risk-bar-fill{{height:100%;border-radius:99px}}
.risk-val{{font-size:12px;font-weight:600;width:42px;text-align:right;flex-shrink:0}}
.chart-wrap{{position:relative;width:100%;height:260px}}
.chart-wrap-sm{{position:relative;width:100%;height:200px}}
.legend-row{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:8px;font-size:12px;color:var(--muted)}}
.legend-item{{display:flex;align-items:center;gap:5px}}
.legend-dot{{width:9px;height:9px;border-radius:2px;flex-shrink:0}}
.countdown{{display:flex;align-items:baseline;gap:3px;margin:2px 0}}
.cd-n{{font-size:30px;font-weight:700;letter-spacing:-1px;line-height:1}}
.cd-u{{font-size:12px;color:var(--muted)}}.cd-sep{{font-size:18px;color:var(--border);margin:0 3px}}
.inp-row{{display:flex;align-items:center;gap:8px;margin-bottom:8px}}
.inp-row label{{font-size:12px;color:var(--muted);width:110px;flex-shrink:0}}
.inp-row input{{flex:1;height:32px;font-size:13px;border:1px solid var(--border);border-radius:var(--radius-sm);padding:0 10px;background:var(--bg);color:var(--text);font-family:inherit}}
.btn-primary{{padding:8px 0;font-size:13px;border-radius:var(--radius-sm);border:none;background:var(--text);cursor:pointer;color:var(--card);font-family:inherit;width:100%;margin-top:8px}}
.badge{{display:inline-flex;align-items:center;padding:2px 9px;border-radius:99px;font-size:11px;font-weight:600}}
.badge-amber{{background:#FAEEDA;color:#633806}}.badge-red{{background:#FCEBEB;color:#501313}}
.holding-header{{font-size:10px;font-weight:600;color:var(--hint);letter-spacing:.06em;text-transform:uppercase;padding:5px 0;border-bottom:1.5px solid var(--border);margin:10px 0 2px}}
.holding-header:first-child{{margin-top:0}}
.sep{{height:1px;background:var(--border);margin:10px 0}}
.total-row{{display:flex;justify-content:space-between;font-size:13px;font-weight:600;color:var(--muted);padding:7px 0;border-top:1px solid var(--border);margin-top:4px}}
.gauge-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:12px}}
.gauge-box{{background:var(--bg);border-radius:12px;padding:12px;display:flex;flex-direction:column;align-items:center}}
.gauge-sub-grid{{display:grid;grid-template-columns:1fr 1fr;gap:8px}}
.gauge-sub-box{{background:var(--bg);border-radius:10px;padding:8px;display:flex;flex-direction:column;align-items:center}}
.gauge-label{{font-size:10px;font-weight:600;letter-spacing:.05em;text-transform:uppercase;color:var(--muted);margin-bottom:4px;text-align:center}}
</style>
</head>
<body>
<div class="wrap">

<div class="header">
  <div>
    <h1>RetireFlow 退休財務儀表板</h1>
    <p>更新時間：{m["update_time"]} ｜ 目標：55歲前退休（{CONFIG["RETIRE_YEAR"]}年前）</p>
  </div>
  <div style="display:flex;gap:8px;flex-wrap:wrap">
    <div class="pill"><span class="dot dot-amber"></span>市場偏謹慎</div>
    <div class="pill"><span class="dot dot-red"></span>台幣升值風險</div>
  </div>
</div>

<div class="g4 mb12">
  <div class="metric">
    <div class="metric-label">總資產</div>
    <div class="metric-value">{total:,}<span style="font-size:13px;font-weight:400;color:var(--muted)">萬</span></div>
    <div class="metric-sub {'up' if m['day_chg']>=0 else 'dn'}">{day_str}</div>
  </div>
  <div class="metric">
    <div class="metric-label">總負債</div>
    <div class="metric-value dn">{debt:,}<span style="font-size:13px;font-weight:400;color:var(--muted)">萬</span></div>
    <div class="metric-sub">負債比 {p(m["debt_ratio"])}%</div>
  </div>
  <div class="metric">
    <div class="metric-label">淨資產</div>
    <div class="metric-value">{net:,}<span style="font-size:13px;font-weight:400;color:var(--muted)">萬</span></div>
    <div class="metric-sub up">↑ 期間 {peri_str}</div>
  </div>
  <div class="metric">
    <div class="metric-label">距退休</div>
    <div class="countdown">
      <span class="cd-n">{m["retire_yrs"]}</span><span class="cd-u">年</span>
      <span class="cd-sep">·</span>
      <span class="cd-n">{m["retire_mths"]}</span><span class="cd-u">月</span>
    </div>
    <div class="metric-sub">年股息 {div}萬｜月均 {round(div/12)}萬</div>
  </div>
</div>

<div class="card mb12">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
    <div>
      <div class="section-title" style="margin-bottom:2px">退休目標達成率</div>
      <div style="font-size:12px;color:var(--muted)">總資產 {total:,}萬 ／ 目標 {RT//10000:,}萬</div>
    </div>
    <div style="text-align:right">
      <div style="font-size:26px;font-weight:700;letter-spacing:-1px">{ach}%</div>
      <div style="font-size:11px;color:var(--muted)">還需增加 {need:,}萬</div>
    </div>
  </div>
  <div class="prog-bg" style="height:12px">
    <div class="prog-fill" style="width:{min(float(ach),100)}%;background:linear-gradient(to right,#378ADD,#1D9E75)"></div>
  </div>
  <div style="display:flex;justify-content:space-between;font-size:11px;color:var(--hint);margin-top:4px">
    <span>起始 {F["date"]}</span><span>目標 {RT//10000:,}萬 ▶</span>
  </div>
</div>

<div class="tab-bar">
  <button class="tab active" onclick="switchTab(this,'overview')">總覽</button>
  <button class="tab" onclick="switchTab(this,'gauge')">油表達成率</button>
  <button class="tab" onclick="switchTab(this,'risk')">風險管控</button>
  <button class="tab" onclick="switchTab(this,'holdings')">持股明細</button>
  <button class="tab" onclick="switchTab(this,'trend')">走勢圖</button>
  <button class="tab" onclick="switchTab(this,'calc')">退休試算</button>
</div>

<!-- 總覽 -->
<div id="tab-overview" class="tab-content active">
  <div class="g2">
    <div class="card">
      <div class="section-title">各類資產現值</div>
      <div class="row"><div><div class="row-name">台股</div><div class="row-tag">永豐・玉山・台銀・凱基</div></div><div style="text-align:right"><div style="font-weight:600">{tw:,}萬</div><div style="font-size:11px;color:#378ADD">{tw_p}% {s(m["tw_ret"])}%</div></div></div>
      <div class="row"><div><div class="row-name">美股</div><div class="row-tag">永豐豐存股・複委託・FirstTrade</div></div><div style="text-align:right"><div style="font-weight:600">{us:,}萬</div><div style="font-size:11px;color:#1D9E75">{us_p}% {s(m["us_ret"])}%</div></div></div>
      <div class="row"><div><div class="row-name">日股</div><div class="row-tag">永豐複委託</div></div><div style="text-align:right"><div style="font-weight:600">{jp:,}萬</div><div style="font-size:11px;color:#EF9F27">{jp_p}% {s(m["jp_ret"])}%</div></div></div>
      <div class="row"><div><div class="row-name">基金</div><div class="row-tag">基富通・鉅亨・HSBC</div></div><div style="text-align:right"><div style="font-weight:600">{fund:,}萬</div><div style="font-size:11px;color:#7F77DD">{fund_p}% {s(m["fund_ret"])}%</div></div></div>
      <div class="row"><div><div class="row-name" style="color:var(--muted)">現金倉位</div><div class="row-tag">SGOV＋台幣現金（待填入）</div></div><div style="font-size:11px;color:var(--hint)">目標10%</div></div>
      <div class="total-row"><span>合計</span><span>{total:,}萬</span></div>
    </div>
    <div class="card">
      <div class="section-title">資產配置</div>
      <div class="chart-wrap-sm"><canvas id="donutChart"></canvas></div>
      <div class="legend-row" style="margin-top:8px;justify-content:center">
        <span class="legend-item"><span class="legend-dot" style="background:#378ADD"></span>台股{tw_p}%</span>
        <span class="legend-item"><span class="legend-dot" style="background:#1D9E75"></span>美股{us_p}%</span>
        <span class="legend-item"><span class="legend-dot" style="background:#EF9F27"></span>日股{jp_p}%</span>
        <span class="legend-item"><span class="legend-dot" style="background:#7F77DD"></span>基金{fund_p}%</span>
      </div>
    </div>
  </div>
  <div class="card mb12">
    <div class="section-title">各資產報酬率（{F["date"]}～{m["update_time"][:10]}，{m["period_days"]}天）</div>
    <div class="g4">
      <div class="metric"><div class="metric-label" style="color:#378ADD">台股</div><div class="metric-value up">{s(m["tw_ret"])}%</div><div class="metric-sub">年化{s(m["tw_ann"])}% ｜ +{w(L["tw"]-F["tw"])}萬</div></div>
      <div class="metric"><div class="metric-label" style="color:#1D9E75">美股</div><div class="metric-value up">{s(m["us_ret"])}%</div><div class="metric-sub">年化{s(m["us_ann"])}% ｜ +{w(L["us"]-F["us"])}萬</div></div>
      <div class="metric"><div class="metric-label" style="color:#EF9F27">日股</div><div class="metric-value up">{s(m["jp_ret"])}%</div><div class="metric-sub">年化{s(m["jp_ann"])}% ｜ +{w(L["jp"]-F["jp"])}萬</div></div>
      <div class="metric"><div class="metric-label" style="color:#7F77DD">基金</div><div class="metric-value up">{s(m["fund_ret"])}%</div><div class="metric-sub">年化{s(m["fund_ann"])}% ｜ +{w(L["fund"]-F["fund"])}萬</div></div>
    </div>
    <div class="alert alert-info" style="margin-top:10px">ℹ️ 年化報酬率為短期數據推算，不代表長期可持續</div>
  </div>
  <div class="g2">
    <div class="card">
      <div class="section-title">負債明細</div>
      {debt_rows}
      <div class="row" style="font-weight:600"><span>總負債</span><span class="dn">{debt:,}萬</span></div>
      <div class="row"><span>年利息</span><span class="dn">約 {int_w}萬</span></div>
      <div class="row"><span>股息淨超過利息</span><span class="{'up' if div_net>=0 else 'dn'}">{'+' if div_net>=0 else ''}{div_net}萬</span></div>
    </div>
    <div class="card">
      <div class="section-title">基金帳戶</div>
      {fund_rows}
      <div class="row" style="font-weight:600"><span>基金小計</span><span>{w(m["fund_total"]):,}萬</span></div>
      <div class="sep"></div>
      <div class="row"><span style="font-size:11px;color:var(--hint)">💡 建議現金倉位達 {w(RT*CONFIG["ALLOC_CASH"]):,}萬（10%）</span></div>
    </div>
  </div>
</div>

<!-- 油表達成率 -->
<div id="tab-gauge" class="tab-content">
  <div class="card mb12">
    <div class="section-title">各類資產退休目標達成率</div>
    <div style="font-size:12px;color:var(--muted)">退休目標 {RT//10000:,}萬，目標配置：台股40%・美股30%・日股10%・基金10%・現金10%</div>
    <div class="gauge-grid">
      <div class="gauge-box">
        <div class="gauge-label">整體達成率</div>
        <canvas id="gauge-overall" width="200" height="110"></canvas>
      </div>
      <div class="gauge-sub-grid">
        <div class="gauge-sub-box"><div class="gauge-label">台股</div><canvas id="gauge-tw" width="130" height="75"></canvas></div>
        <div class="gauge-sub-box"><div class="gauge-label">美股</div><canvas id="gauge-us" width="130" height="75"></canvas></div>
        <div class="gauge-sub-box"><div class="gauge-label">日股</div><canvas id="gauge-jp" width="130" height="75"></canvas></div>
        <div class="gauge-sub-box"><div class="gauge-label">基金</div><canvas id="gauge-fund" width="130" height="75"></canvas></div>
      </div>
    </div>
  </div>
  <div class="card mb12">
    <div class="section-title">目標配置 vs 實際配置</div>
    <div style="margin-top:8px">
      <div class="risk-row"><div class="risk-label" style="color:#378ADD">台股</div><div style="flex:1"><div style="display:flex;justify-content:space-between;font-size:11px;color:var(--muted);margin-bottom:3px"><span>實際 {tw_p}%</span><span>目標 40%</span></div><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:{min(m["tw_pct"]/40*100,100):.0f}%;background:#378ADD"></div></div></div><div class="risk-val" style="color:{tw_ach_color}">{tw_ach_diff}%</div></div>
      <div class="risk-row"><div class="risk-label" style="color:#1D9E75">美股</div><div style="flex:1"><div style="display:flex;justify-content:space-between;font-size:11px;color:var(--muted);margin-bottom:3px"><span>實際 {us_p}%</span><span>目標 30%</span></div><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:{min(m["us_pct"]/30*100,100):.0f}%;background:#1D9E75"></div></div></div><div class="risk-val" style="color:#1D9E75">{us_ach_diff}%</div></div>
      <div class="risk-row"><div class="risk-label" style="color:#EF9F27">日股</div><div style="flex:1"><div style="display:flex;justify-content:space-between;font-size:11px;color:var(--muted);margin-bottom:3px"><span>實際 {jp_p}%</span><span>目標 10%</span></div><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:{min(m["jp_pct"]/10*100,100):.0f}%;background:#EF9F27"></div></div></div><div class="risk-val" style="color:#1D9E75">{jp_ach_diff}%</div></div>
      <div class="risk-row"><div class="risk-label" style="color:#7F77DD">基金</div><div style="flex:1"><div style="display:flex;justify-content:space-between;font-size:11px;color:var(--muted);margin-bottom:3px"><span>實際 {fund_p}%</span><span>目標 10%</span></div><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:{min(m["fund_pct"]/10*100,100):.0f}%;background:#7F77DD"></div></div></div><div class="risk-val" style="color:#1D9E75">{fd_ach_diff}%</div></div>
      <div class="risk-row"><div class="risk-label" style="color:var(--muted)">現金</div><div style="flex:1"><div style="display:flex;justify-content:space-between;font-size:11px;color:var(--muted);margin-bottom:3px"><span>實際 ~0%</span><span>目標 10%</span></div><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:3%;background:#999"></div></div></div><div class="risk-val dn">-10%</div></div>
    </div>
  </div>
  <div class="alert alert-danger">⚠️ <b>台股超重 {tw_ach_diff}%</b>：建議分批降低槓桿ETF部位，轉入現金或美股</div>
  <div class="alert alert-danger" style="margin-top:0">⚠️ <b>現金倉位不足</b>：建議增至 {w(RT*CONFIG["ALLOC_CASH"]):,}萬，作為回調加碼彈藥</div>
  <div class="alert alert-good" style="margin-top:8px">✓ 美股・日股・基金配置偏差在合理範圍（±3%）</div>
</div>

<!-- 風險管控 -->
<div id="tab-risk" class="tab-content">
  <div class="g3 mb12">
    <div class="card">
      <div class="section-title">CNN 恐懼貪婪</div>
      <div style="display:flex;align-items:baseline;gap:8px;margin:8px 0"><span style="font-size:34px;font-weight:700;color:#EF9F27">41</span><span class="badge badge-amber">恐懼</span></div>
      <div class="vix-bar"><div class="vix-pin" style="left:41%"></div></div>
      <div style="display:flex;justify-content:space-between;font-size:10px;color:var(--hint)"><span>極恐懼</span><span>中性50</span><span>極貪婪</span></div>
      <div class="alert alert-warn" style="margin-top:10px">逢低分批佈局，避免重押</div>
    </div>
    <div class="card">
      <div class="section-title">VIX 波動率</div>
      <div style="display:flex;align-items:baseline;gap:8px;margin:8px 0"><span style="font-size:34px;font-weight:700;color:#E24B4A">22.3</span><span class="badge badge-red">警戒</span></div>
      <div class="vix-bar" style="background:linear-gradient(to right,#1D9E75 40%,#EF9F27 65%,#E24B4A)"><div class="vix-pin" style="left:55%"></div></div>
      <div style="display:flex;justify-content:space-between;font-size:10px;color:var(--hint)"><span>&lt;15低波</span><span>20警戒</span><span>&gt;30恐慌</span></div>
      <div class="alert alert-danger" style="margin-top:10px">VIX&gt;20 降低槓桿</div>
    </div>
    <div class="card">
      <div class="section-title">市場燈號</div>
      <div class="row"><span>台股加權</span><span><span class="dot dot-green" style="margin-right:5px"></span>多頭</span></div>
      <div class="row"><span>S&amp;P 500</span><span><span class="dot dot-amber" style="margin-right:5px"></span>盤整</span></div>
      <div class="row"><span>Nasdaq</span><span><span class="dot dot-amber" style="margin-right:5px"></span>盤整</span></div>
      <div class="row"><span>日經225</span><span><span class="dot dot-amber" style="margin-right:5px"></span>盤整</span></div>
      <div class="row"><span>USD/TWD</span><span><span class="dot dot-red" style="margin-right:5px"></span>台幣強升</span></div>
      <div class="row"><span>槓桿ETF</span><span><span class="dot dot-red" style="margin-right:5px"></span>耗損風險</span></div>
    </div>
  </div>
  <div class="card mb12">
    <div class="section-title">風險部位評估</div>
    <div style="margin-top:6px">
      <div class="risk-row"><div class="risk-label">台股槓桿</div><div style="flex:1"><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:78%;background:#E24B4A"></div></div></div><div class="risk-val dn">高</div></div>
      <div style="font-size:11px;color:var(--muted);margin:-4px 0 10px 90px">00631L・00663L・00675L 共33,000股</div>
      <div class="risk-row"><div class="risk-label">美股槓桿</div><div style="flex:1"><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:55%;background:#EF9F27"></div></div></div><div class="risk-val" style="color:#EF9F27">中</div></div>
      <div style="font-size:11px;color:var(--muted);margin:-4px 0 10px 90px">TQQQ・QLD・SSO・AMDL</div>
      <div class="risk-row"><div class="risk-label">股票質借</div><div style="flex:1"><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:40%;background:#EF9F27"></div></div></div><div class="risk-val" style="color:#EF9F27">中</div></div>
      <div style="font-size:11px;color:var(--muted);margin:-4px 0 10px 90px">永豐台股 {w(7978550)}萬 @2.5%</div>
      <div class="risk-row"><div class="risk-label">匯率風險</div><div style="flex:1"><div class="risk-bar-bg"><div class="risk-bar-fill" style="width:65%;background:#E24B4A"></div></div></div><div class="risk-val dn">高</div></div>
      <div style="font-size:11px;color:var(--muted);margin:-4px 0 0 90px">美股+日股 {us+jp:,}萬，台幣每升1%損失約 {round((us+jp)*0.01)}萬</div>
    </div>
  </div>
  <div class="alert alert-danger">⚠️ <b>台幣快速升值</b>：外幣資產 {us+jp:,}萬，匯率風險持續升高</div>
  <div class="alert alert-warn" style="margin-top:0">⚠️ <b>槓桿ETF耗損</b>：VIX&gt;20 時 volatility decay 明顯，逢反彈減碼</div>
  <div class="alert alert-good" style="margin-top:8px">✓ <b>年股息{div}萬 &gt; 年利息{int_w}萬</b>，財務正向循環</div>
  <div class="alert alert-good" style="margin-top:0">✓ 高股息除息旺季（Q2-Q3）即將到來</div>
</div>

<!-- 持股明細 -->
<div id="tab-holdings" class="tab-content">
  <div class="g2">
    <div class="card"><div class="section-title">台股持股</div>{tw_rows}</div>
    <div class="card">
      <div class="section-title">美股 / 日股</div>
      <div class="holding-header">美股</div>{us_rows}
      <div class="holding-header">日股（永豐複委託）</div>{jp_rows}
    </div>
  </div>
</div>

<!-- 走勢圖 -->
<div id="tab-trend" class="tab-content">
  <div class="card mb12">
    <div class="section-title">總資產 / 淨資產 / 負債</div>
    <div class="legend-row"><span class="legend-item"><span class="legend-dot" style="background:#378ADD"></span>總資產</span><span class="legend-item"><span class="legend-dot" style="background:#1D9E75"></span>淨資產</span><span class="legend-item"><span class="legend-dot" style="background:#E24B4A"></span>總負債</span></div>
    <div class="chart-wrap"><canvas id="trendChart"></canvas></div>
  </div>
  <div class="card mb12">
    <div class="section-title">各類資產走勢</div>
    <div class="legend-row"><span class="legend-item"><span class="legend-dot" style="background:#378ADD"></span>台股</span><span class="legend-item"><span class="legend-dot" style="background:#1D9E75"></span>美股</span><span class="legend-item"><span class="legend-dot" style="background:#EF9F27"></span>日股</span><span class="legend-item"><span class="legend-dot" style="background:#7F77DD"></span>基金</span></div>
    <div class="chart-wrap"><canvas id="allocChart"></canvas></div>
  </div>
  <div class="card">
    <div class="section-title">每日資產變化</div>
    <div class="chart-wrap" style="height:180px"><canvas id="barChart"></canvas></div>
  </div>
</div>

<!-- 退休試算 -->
<div id="tab-calc" class="tab-content">
  <div class="g2">
    <div class="card">
      <div class="section-title">退休目標設定</div>
      <div class="inp-row"><label>退休目標（萬）</label><input type="number" id="s-target" value="{RT//10000}" step="100"></div>
      <div class="inp-row"><label>目前總資產（萬）</label><input type="number" id="s-current" value="{total}"></div>
      <div class="inp-row"><label>每月新增（萬）</label><input type="number" id="s-monthly" value="15"></div>
      <div class="inp-row"><label>預期年報酬率（%）</label><input type="number" id="s-return" value="10" step="0.5"></div>
      <div class="inp-row"><label>退休後月支出（萬）</label><input type="number" id="s-expense" value="12"></div>
      <div class="inp-row"><label>目前年齡</label><input type="number" id="s-age" value="{CONFIG["CURRENT_AGE"]}"></div>
      <button class="btn-primary" onclick="calcRetire()">重新計算 →</button>
    </div>
    <div class="card">
      <div class="section-title">試算結果</div>
      <div style="margin-top:8px">
        <div class="metric mb12"><div class="metric-label">預計達到退休門檻</div><div class="metric-value" id="r-years">—</div><div class="metric-sub" id="r-date"></div></div>
        <div class="metric mb12"><div class="metric-label">退休時預估總資產</div><div class="metric-value" id="r-final">—</div><div class="metric-sub" id="r-fs"></div></div>
        <div class="metric"><div class="metric-label">退休金可支應年限</div><div class="metric-value" id="r-dur">—</div><div class="metric-sub">4% 提領法則</div></div>
      </div>
    </div>
  </div>
  <div class="card" style="margin-top:12px">
    <div class="section-title">財務摘要</div>
    <div class="row"><span>期間報酬（{m["period_days"]}天）</span><span class="up" style="font-weight:600">{s(m["period_pct"])}%</span></div>
    <div class="row"><span>預估年股息</span><span>{div}萬</span></div>
    <div class="row"><span>年利息支出</span><span class="dn">{int_w}萬</span></div>
    <div class="row"><span>股息淨超過利息</span><span class="up">+{div_net}萬</span></div>
    <div class="row"><span>月均被動收入</span><span class="up">{round(div/12)}萬</span></div>
  </div>
</div>

<div style="text-align:center;font-size:11px;color:var(--hint);margin-top:20px;padding-top:14px;border-top:1px solid var(--border)">
  RetireFlow v2.0 ｜ {m["update_time"]} 自動產生 ｜ 僅供個人參考，不構成投資建議
</div>
</div>

<script>
const H={{dates:{h_dates},total:{h_total},net:{h_net},debt:{h_debt},tw:{h_tw},us:{h_us},jp:{h_jp},fund:{h_fund}}};
const dark=matchMedia('(prefers-color-scheme:dark)').matches;
const gc=dark?'rgba(255,255,255,0.07)':'rgba(0,0,0,0.06)';
const tc=dark?'#555':'#aaa';

function drawGauge(id,pct,label){{
  const c=document.getElementById(id);if(!c)return;
  const W=c.clientWidth||c.width,H2=c.clientHeight||c.height;
  c.width=W;c.height=H2;
  const ctx=c.getContext('2d');
  const cx=W/2,cy=H2*0.88,r=Math.min(W,H2*2)*0.38;
  const cl=Math.min(pct,100);
  const col=pct<40?'#E24B4A':pct<70?'#EF9F27':pct<90?'#1D9E75':'#378ADD';
  ctx.beginPath();ctx.arc(cx,cy,r,Math.PI,2*Math.PI);
  ctx.strokeStyle=dark?'#333':'#eee';ctx.lineWidth=r*0.22;ctx.lineCap='round';ctx.stroke();
  if(cl>0){{ctx.beginPath();ctx.arc(cx,cy,r,Math.PI,Math.PI+Math.PI*cl/100);
  ctx.strokeStyle=col;ctx.lineWidth=r*0.22;ctx.lineCap='round';ctx.stroke();}}
  ctx.fillStyle=dark?'#f0efe9':'#1a1a1a';
  ctx.font=`bold ${{r*0.5}}px -apple-system,sans-serif`;
  ctx.textAlign='center';ctx.textBaseline='middle';
  ctx.fillText(pct.toFixed(1)+'%',cx,cy-r*0.12);
  ctx.fillStyle=tc;ctx.font=`${{r*0.28}}px -apple-system,sans-serif`;
  ctx.fillText(label||'',cx,cy+r*0.3);
}}

let gaugesDrawn=false;
function drawAllGauges(){{
  if(gaugesDrawn)return;gaugesDrawn=true;
  drawGauge('gauge-overall',{float(ach)},'{total:,}萬 / {RT//10000:,}萬');
  drawGauge('gauge-tw',{float(tw_a)},'{tw:,}萬');
  drawGauge('gauge-us',{float(us_a)},'{us:,}萬');
  drawGauge('gauge-jp',{float(jp_a)},'{jp:,}萬');
  drawGauge('gauge-fund',{float(fund_a)},'{fund:,}萬');
}}

function mkLine(id,datasets){{
  const ctx=document.getElementById(id);if(!ctx)return;
  new Chart(ctx,{{type:'line',data:{{labels:H.dates,datasets}},options:{{responsive:true,maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}},plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>` ${{c.dataset.label}}: ${{c.parsed.y.toLocaleString()}}萬`}}}}}},scales:{{x:{{grid:{{color:gc}},ticks:{{color:tc,font:{{size:10}},maxTicksLimit:10}}}},y:{{grid:{{color:gc}},ticks:{{color:tc,font:{{size:10}},callback:v=>v+'萬'}}}}}}}}}});
}}

let chartsDrawn=false;
function drawCharts(){{
  if(chartsDrawn)return;chartsDrawn=true;
  mkLine('trendChart',[{{label:'總資產',data:H.total,borderColor:'#378ADD',borderWidth:2,pointRadius:0,fill:true,backgroundColor:dark?'rgba(55,138,221,0.08)':'rgba(55,138,221,0.07)',tension:0.3}},{{label:'淨資產',data:H.net,borderColor:'#1D9E75',borderWidth:2,pointRadius:0,fill:false,tension:0.3}},{{label:'總負債',data:H.debt,borderColor:'#E24B4A',borderWidth:1.5,borderDash:[4,3],pointRadius:0,fill:false,tension:0.3}}]);
  mkLine('allocChart',[{{label:'台股',data:H.tw,borderColor:'#378ADD',borderWidth:2,pointRadius:0,fill:false,tension:0.3}},{{label:'美股',data:H.us,borderColor:'#1D9E75',borderWidth:2,pointRadius:0,fill:false,tension:0.3}},{{label:'日股',data:H.jp,borderColor:'#EF9F27',borderWidth:2,pointRadius:0,fill:false,tension:0.3}},{{label:'基金',data:H.fund,borderColor:'#7F77DD',borderWidth:2,pointRadius:0,fill:false,tension:0.3}}]);
  const chg=H.total.map((v,i)=>i===0?0:v-H.total[i-1]);
  const bctx=document.getElementById('barChart');
  if(bctx) new Chart(bctx,{{type:'bar',data:{{labels:H.dates,datasets:[{{label:'每日變化',data:chg,backgroundColor:chg.map(v=>v>=0?'rgba(29,158,117,0.75)':'rgba(226,75,74,0.75)'),borderRadius:3}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>` ${{c.parsed.y>=0?'+':''}}${{c.parsed.y}}萬`}}}}}},scales:{{x:{{grid:{{display:false}},ticks:{{color:tc,font:{{size:10}},maxTicksLimit:10}}}},y:{{grid:{{color:gc}},ticks:{{color:tc,font:{{size:10}},callback:v=>v+'萬'}}}}}}}}}});
  new Chart(document.getElementById('donutChart'),{{type:'doughnut',data:{{labels:['台股','美股','日股','基金'],datasets:[{{data:[{tw_p},{us_p},{jp_p},{fund_p}],backgroundColor:['#378ADD','#1D9E75','#EF9F27','#7F77DD'],borderWidth:0,hoverOffset:4}}]}},options:{{responsive:true,maintainAspectRatio:false,cutout:'66%',plugins:{{legend:{{display:false}}}}}}}});
}}

function switchTab(el,name){{
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab-content').forEach(t=>t.classList.remove('active'));
  el.classList.add('active');document.getElementById('tab-'+name).classList.add('active');
  if(name==='trend')setTimeout(drawCharts,60);
  if(name==='gauge')setTimeout(drawAllGauges,60);
}}

function calcRetire(){{
  const tgt=parseFloat(document.getElementById('s-target').value)||{RT//10000};
  const cur=parseFloat(document.getElementById('s-current').value)||{total};
  const mon=parseFloat(document.getElementById('s-monthly').value)||15;
  const ret=parseFloat(document.getElementById('s-return').value)/100;
  const exp=parseFloat(document.getElementById('s-expense').value)||12;
  const age0=parseFloat(document.getElementById('s-age').value)||{CONFIG["CURRENT_AGE"]};
  const mr=ret/12;let bal=cur,months=0;
  while(bal<tgt&&months<600){{bal=bal*(1+mr)+mon;months++;}}
  const yr=2026+Math.floor(months/12);
  const age=(age0+months/12).toFixed(1);
  const dur=Math.min((bal*0.04/(exp*12)*25),99).toFixed(1);
  document.getElementById('r-years').textContent=(months/12).toFixed(1)+'年後';
  document.getElementById('r-date').textContent='約'+yr+'年（'+age+'歲）';
  document.getElementById('r-final').textContent='NT$'+Math.round(bal).toLocaleString()+'萬';
  document.getElementById('r-fs').textContent='目標'+tgt.toLocaleString()+'萬 達成';
  document.getElementById('r-dur').textContent=dur+'年';
}}

document.addEventListener('DOMContentLoaded',()=>{{drawCharts();drawAllGauges();calcRetire();}});
</script>
</body>
</html>"""
    return html

# ── 主程式 ────────────────────────────────────────────────────
def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "RetireFlow_DB.xlsx"
    output = Path("index.html")
    print(f"📂 讀取 {xlsx_path} ...")
    history, debts, funds, holdings = read_data(xlsx_path)
    print(f"✅ 歷史紀錄 {len(history)} 筆，持股 台股{len(holdings['tw'])} 美股{len(holdings['us'])} 日股{len(holdings['jp'])}")
    metrics = calc_metrics(history, debts, funds, holdings)
    html = build_html(metrics)
    output.write_text(html, encoding="utf-8")
    print(f"✅ 儀表板已產生：{output.resolve()}")
    print(f"   總資產：{w(metrics['L']['total']):,}萬  達成率：{p(metrics['achieve'])}%")

if __name__ == "__main__":
    main()
